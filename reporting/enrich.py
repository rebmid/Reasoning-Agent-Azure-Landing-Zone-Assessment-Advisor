"""Post-processing enrichment for the Checklist sheet.

Normalises every row so that it has:
  - a Control ID
  - a valid ALZ Design Area
  - a WAF pillar
  - traceability metadata (Control Source, Derived Control ID, Control Type,
    Related ALZ Control IDs)

This runs **after** the workbook is already written — it never touches
evaluator logic, scoring, roadmap generation, or the ALZ fetch pipeline.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# ══════════════════════════════════════════════════════════════════
# Constants
# ══════════════════════════════════════════════════════════════════

_SHEET_NAME = "Checklist"
_HEADER_ROW = 9
_DATA_START = 10

# Existing column layout (A-U = 1-21)
_COL_ID = 1           # A  e.g. "D07.01"
_COL_DESIGN_AREA = 2  # B
_COL_SUB_AREA = 3     # C
_COL_WAF = 4          # D
_COL_SERVICE = 5      # E
_COL_TEXT = 6          # F  Checklist item
_COL_SEVERITY = 8     # H
_COL_STATUS = 9       # I
_COL_GUID = 15        # O

# New enrichment columns (appended after U=21)
_COL_CONTROL_SOURCE = 22      # V
_COL_DERIVED_CONTROL_ID = 23  # W
_COL_CONTROL_TYPE = 24        # X
_COL_RELATED_ALZ = 25         # Y

_NEW_HEADERS = {
    _COL_CONTROL_SOURCE: "Control Source",
    _COL_DERIVED_CONTROL_ID: "Derived Control ID",
    _COL_CONTROL_TYPE: "Control Type",
    _COL_RELATED_ALZ: "Related ALZ Control IDs",
}

# ALZ ID pattern: <letter><digits>.<digits>  (e.g. D01.01, C02.03)
_ALZ_ID_RE = re.compile(r"^[A-Z]\d{2}\.\d{2}$")


# ══════════════════════════════════════════════════════════════════
# Design Area keyword mapping  (STEP 5)
# ══════════════════════════════════════════════════════════════════

_DESIGN_AREA_KEYWORDS: list[tuple[list[str], str]] = [
    (["firewall", "vnet", "virtual network", "nsg", "ddos", "private endpoint",
      "dns", "expressroute", "vpn", "front door", "application gateway",
      "load balancer", "network watcher", "aks"],
     "Network Topology and Connectivity"),

    (["rbac", "pim", "entra", "identity", "conditional access",
      "break-glass", "breakglass", "service principal"],
     "Identity and Access Management"),

    (["key vault", "keyvault", "defender", "sql posture", "storage posture",
      "app service", "acr", "container registry", "encryption",
      "private endpoint"],
     "Security"),

    (["backup", "restore", "disaster recovery", "availability",
      "resilience", "slo"],
     "Management and Monitoring"),

    (["diagnostic", "log", "monitor", "alert", "sentinel", "workspace",
      "update manager", "change tracking", "action group", "activity log"],
     "Management and Monitoring"),

    (["management group", "resource lock", "policy", "tag", "naming",
      "subscription", "governance"],
     "Governance"),

    (["cost", "budget", "reservation", "idle resource", "forecast"],
     "Cost Management"),

    (["platform automation", "devops", "iac", "bicep", "terraform"],
     "Platform Automation and DevOps"),
]


def _infer_design_area(text: str) -> str:
    """Infer design area from control text using keyword matching."""
    lower = text.lower()
    for keywords, area in _DESIGN_AREA_KEYWORDS:
        if any(kw in lower for kw in keywords):
            return area
    return "Security"  # default fallback


# ══════════════════════════════════════════════════════════════════
# WAF pillar mapping  (STEP 6)
# ══════════════════════════════════════════════════════════════════

_DESIGN_AREA_TO_WAF: dict[str, str] = {
    "Identity and Access Management":       "Security",
    "Security":                             "Security",
    "Network Topology and Connectivity":    "Security",
    "Management and Monitoring":            "Operational Excellence",
    "Management":                           "Operational Excellence",
    "Governance":                           "Operational Excellence",
    "Resource Organization":                "Operational Excellence",
    "Platform Automation and DevOps":       "Operational Excellence",
    "Cost Management":                      "Cost Optimization",
    "Cost Optimization":                    "Cost Optimization",
    # Fallback for ALZ legacy names
    "Azure Billing and Microsoft Entra ID Tenants": "Security",
}


def _infer_waf(design_area: str) -> str:
    """Map design area to WAF pillar."""
    return _DESIGN_AREA_TO_WAF.get(design_area, "Security")


# ══════════════════════════════════════════════════════════════════
# Deterministic derived control ID  (STEP 4)
# ══════════════════════════════════════════════════════════════════

_AREA_SLUG: dict[str, str] = {
    "Network Topology and Connectivity":    "NET",
    "Identity and Access Management":       "ID",
    "Security":                             "SEC",
    "Management and Monitoring":            "MG",
    "Management":                           "MG",
    "Governance":                           "GOV",
    "Cost Management":                      "COST",
    "Cost Optimization":                    "COST",
    "Platform Automation and DevOps":       "AUTO",
    "Resource Organization":                "ORG",
    "Data Protection":                      "SEC",
    "Resilience":                           "RES",
    "Networking":                           "NET",
}


def _make_derived_id(design_area: str, text: str) -> str:
    """Generate DER-<AREA>-<SLUG> deterministic ID."""
    area_code = _AREA_SLUG.get(design_area, "GEN")
    # Uppercase, replace non-alphanumeric with dash, collapse, trim
    slug = re.sub(r"[^A-Z0-9]+", "-", text.upper().strip())
    slug = slug.strip("-")[:40]
    return f"DER-{area_code}-{slug}"


# ══════════════════════════════════════════════════════════════════
# Related ALZ control matching  (STEP 8)
# ══════════════════════════════════════════════════════════════════

# Keyword groups for cross-referencing derived → ALZ
_TRACEABILITY_KEYWORDS: dict[str, list[str]] = {
    "storage":         ["storage", "diagnostics", "encryption", "defender"],
    "keyvault":        ["key vault", "keyvault", "encryption", "secrets"],
    "sql":             ["sql", "database", "defender", "audit"],
    "appservice":      ["app service", "web app", "defender"],
    "private-endpoint": ["private endpoint", "private link", "network"],
    "acr":             ["container registry", "acr", "defender"],
    "nsg":             ["nsg", "network security group", "subnet"],
    "aks":             ["aks", "kubernetes", "container"],
    "backup":          ["backup", "recovery", "vault"],
    "resource-lock":   ["resource lock", "lock", "protect"],
    "rbac":            ["rbac", "role", "access", "privilege"],
    "entra":           ["entra", "azure ad", "identity", "sign-in"],
    "pim":             ["pim", "privileged", "just-in-time"],
    "monitor":         ["monitor", "log analytics", "workspace", "sentinel"],
    "diagnostic":      ["diagnostic", "activity log", "audit log"],
    "alert":           ["alert", "action group", "notification"],
    "update":          ["update", "patch", "change tracking"],
    "cost":            ["cost", "budget", "spending", "reservation"],
    "firewall":        ["firewall", "hub", "spoke", "egress"],
    "defender":        ["defender", "security center", "secure score"],
    "network-watcher": ["network watcher", "flow log", "observability"],
    "idle":            ["idle", "unused", "orphan"],
}


def _find_related_alz(derived_text: str, alz_rows: list[dict[str, str]]) -> str:
    """Find ALZ control IDs whose text overlaps with derived control keywords.

    Returns semicolon-separated list of matching ALZ IDs (max 5).
    """
    lower = derived_text.lower()
    # Collect all matching keyword families
    relevant_kws: set[str] = set()
    for _group, kws in _TRACEABILITY_KEYWORDS.items():
        if any(kw in lower for kw in kws):
            relevant_kws.update(kws)

    if not relevant_kws:
        return ""

    matches: list[str] = []
    for alz in alz_rows:
        alz_text = alz.get("text", "").lower()
        if any(kw in alz_text for kw in relevant_kws):
            alz_id = alz.get("id", "")
            if alz_id and alz_id not in matches:
                matches.append(alz_id)
                if len(matches) >= 5:
                    break

    return "; ".join(matches)


# ══════════════════════════════════════════════════════════════════
# Main enrichment function
# ══════════════════════════════════════════════════════════════════

def enrich_control_details_sheet(workbook_path: str) -> dict[str, Any]:
    """Post-process the Checklist sheet to normalise all rows.

    This is called AFTER the workbook is written.  It adds 4 columns
    (V-Y) and populates metadata for both ALZ and derived controls.

    Returns a validation summary dict.
    """
    wb = load_workbook(workbook_path, keep_vba=True)
    if _SHEET_NAME not in wb.sheetnames:
        wb.close()
        return {"error": f"Sheet '{_SHEET_NAME}' not found"}

    ws = wb[_SHEET_NAME]

    # ── STEP 2: Write new column headers ──────────────────────────
    header_font = Font(bold=True, color="FFFFFF")
    from openpyxl.styles import PatternFill
    header_fill = PatternFill("solid", fgColor="4472C4")
    for col_idx, header_text in _NEW_HEADERS.items():
        cell = ws.cell(row=_HEADER_ROW, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill

    # ── Collect all rows ──────────────────────────────────────────
    last_row = ws.max_row
    if last_row < _DATA_START:
        wb.save(workbook_path)
        wb.close()
        return {"rows_processed": 0, "alz": 0, "derived": 0}

    # First pass: collect ALZ rows for traceability linking
    alz_rows_for_linking: list[dict[str, str]] = []
    for row in range(_DATA_START, last_row + 1):
        item_id = str(ws.cell(row=row, column=_COL_ID).value or "")
        if _ALZ_ID_RE.match(item_id):
            alz_rows_for_linking.append({
                "id": item_id,
                "text": str(ws.cell(row=row, column=_COL_TEXT).value or ""),
                "design_area": str(ws.cell(row=row, column=_COL_DESIGN_AREA).value or ""),
            })

    # ── Second pass: enrich every row ─────────────────────────────
    stats = {"rows_processed": 0, "alz": 0, "derived": 0,
             "design_area_filled": 0, "waf_filled": 0}

    for row in range(_DATA_START, last_row + 1):
        item_id = str(ws.cell(row=row, column=_COL_ID).value or "").strip()
        text = str(ws.cell(row=row, column=_COL_TEXT).value or "").strip()
        design_area = str(ws.cell(row=row, column=_COL_DESIGN_AREA).value or "").strip()
        waf = str(ws.cell(row=row, column=_COL_WAF).value or "").strip()

        # Skip truly empty rows
        if not item_id and not text:
            continue

        stats["rows_processed"] += 1
        is_alz = bool(_ALZ_ID_RE.match(item_id))

        # ── STEP 1 & 3: Control Source ────────────────────────────
        if is_alz:
            stats["alz"] += 1
            ws.cell(row=row, column=_COL_CONTROL_SOURCE, value="ALZ")
            ws.cell(row=row, column=_COL_DERIVED_CONTROL_ID, value="")
            ws.cell(row=row, column=_COL_CONTROL_TYPE, value="Foundational")
            ws.cell(row=row, column=_COL_RELATED_ALZ, value="")
        else:
            stats["derived"] += 1
            ws.cell(row=row, column=_COL_CONTROL_SOURCE, value="Derived")

            # ── STEP 5: Design Area auto-mapping ──────────────────
            if not design_area:
                design_area = _infer_design_area(text)
                ws.cell(row=row, column=_COL_DESIGN_AREA, value=design_area)
                stats["design_area_filled"] += 1

            # ── STEP 6: WAF pillar mapping ────────────────────────
            if not waf:
                waf = _infer_waf(design_area)
                ws.cell(row=row, column=_COL_WAF, value=waf)
                stats["waf_filled"] += 1

            # ── STEP 4: Deterministic derived ID ──────────────────
            derived_id = _make_derived_id(design_area, text)
            ws.cell(row=row, column=_COL_DERIVED_CONTROL_ID, value=derived_id)

            # ── STEP 7: Control Type ──────────────────────────────
            ws.cell(row=row, column=_COL_CONTROL_TYPE, value="Detective")

            # ── STEP 8: Related ALZ Control IDs ───────────────────
            related = _find_related_alz(text, alz_rows_for_linking)
            ws.cell(row=row, column=_COL_RELATED_ALZ, value=related)

    # ── Save ──────────────────────────────────────────────────────
    try:
        wb.save(workbook_path)
    except PermissionError:
        from datetime import datetime
        p = Path(workbook_path)
        ts = datetime.now().strftime("%H%M%S")
        fallback = p.with_name(f"{p.stem}_enriched_{ts}.xlsm")
        wb.save(str(fallback))
        print(f"  ⚠ Saved enriched workbook as {fallback.name} (original locked)")
    wb.close()

    # ── STEP 10: Validation ───────────────────────────────────────
    _validate(stats)

    return stats


def _validate(stats: dict[str, Any]) -> None:
    """Print validation summary."""
    total = stats["rows_processed"]
    alz = stats["alz"]
    derived = stats["derived"]
    print(f"  ✓ Enrichment: {total} rows processed — {alz} ALZ, {derived} Derived")
    if stats["design_area_filled"]:
        print(f"    → {stats['design_area_filled']} derived rows: Design Area auto-mapped")
    if stats["waf_filled"]:
        print(f"    → {stats['waf_filled']} derived rows: WAF pillar auto-mapped")
    if derived > 0:
        print(f"    → All derived rows have: Derived Control ID ✅, Control Source ✅")
    print(f"    → All rows have: Control Source ✅, Control Type ✅")


# ══════════════════════════════════════════════════════════════════
# Public API — enrich an already-open worksheet (no file I/O)
# ══════════════════════════════════════════════════════════════════

def enrich_open_worksheet(ws) -> dict[str, Any]:
    """Apply enrichment columns (V–Y) to an already-open worksheet.

    This avoids a second openpyxl load/save cycle.  Called by the main
    workbook builder after writing Checklist data rows.

    Returns a stats dict with counts.
    """
    # ── Write enrichment headers ──────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    for col_idx, header_text in _NEW_HEADERS.items():
        cell = ws.cell(row=_HEADER_ROW, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill

    last_row = ws.max_row
    if last_row < _DATA_START:
        return {"rows_processed": 0, "alz": 0, "derived": 0}

    # ── First pass: collect ALZ rows for traceability ─────────────
    alz_rows_for_linking: list[dict[str, str]] = []
    for row in range(_DATA_START, last_row + 1):
        item_id = str(ws.cell(row=row, column=_COL_ID).value or "")
        if _ALZ_ID_RE.match(item_id):
            alz_rows_for_linking.append({
                "id": item_id,
                "text": str(ws.cell(row=row, column=_COL_TEXT).value or ""),
                "design_area": str(
                    ws.cell(row=row, column=_COL_DESIGN_AREA).value or ""
                ),
            })

    # ── Second pass: enrich every row ─────────────────────────────
    stats: dict[str, Any] = {
        "rows_processed": 0, "alz": 0, "derived": 0,
        "design_area_filled": 0, "waf_filled": 0,
    }

    for row in range(_DATA_START, last_row + 1):
        item_id = str(ws.cell(row=row, column=_COL_ID).value or "").strip()
        text = str(ws.cell(row=row, column=_COL_TEXT).value or "").strip()
        design_area = str(
            ws.cell(row=row, column=_COL_DESIGN_AREA).value or ""
        ).strip()
        waf = str(ws.cell(row=row, column=_COL_WAF).value or "").strip()

        if not item_id and not text:
            continue

        stats["rows_processed"] += 1
        is_alz = bool(_ALZ_ID_RE.match(item_id))

        if is_alz:
            stats["alz"] += 1
            ws.cell(row=row, column=_COL_CONTROL_SOURCE, value="ALZ")
            ws.cell(row=row, column=_COL_DERIVED_CONTROL_ID, value="")
            ws.cell(row=row, column=_COL_CONTROL_TYPE, value="Foundational")
            ws.cell(row=row, column=_COL_RELATED_ALZ, value="")
        else:
            stats["derived"] += 1
            ws.cell(row=row, column=_COL_CONTROL_SOURCE, value="Derived")

            if not design_area:
                design_area = _infer_design_area(text)
                ws.cell(row=row, column=_COL_DESIGN_AREA, value=design_area)
                stats["design_area_filled"] += 1

            if not waf:
                waf = _infer_waf(design_area)
                ws.cell(row=row, column=_COL_WAF, value=waf)
                stats["waf_filled"] += 1

            ws.cell(
                row=row, column=_COL_DERIVED_CONTROL_ID,
                value=_make_derived_id(design_area, text),
            )
            ws.cell(row=row, column=_COL_CONTROL_TYPE, value="Detective")
            ws.cell(
                row=row, column=_COL_RELATED_ALZ,
                value=_find_related_alz(text, alz_rows_for_linking),
            )

    _validate(stats)
    return stats
