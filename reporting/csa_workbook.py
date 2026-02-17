"""CSA Workbook builder — template-based, data-only writer.

Copies the pre-built ``.xlsm`` template and writes **only** data values
into the existing sheets:

- ``0_Executive_Summary`` — engagement framing + assessment metrics + top risks
- ``1_30-60-90_Roadmap`` — phased remediation initiatives
- ``2_Control_Details`` — one row per assessed control (columns A–U)
  plus enrichment metadata (columns V–Y)
- ``3_Risk_Analysis`` — causal risk blocks with failing controls,
  dependency impact, remediation roadmap, and cascade effect

The template owns **all** visualisation: Dashboard formulas, charts,
conditional formatting, data validation, and VBA macros.  Python never
creates, modifies, or deletes sheets, formatting, formulas, or macros.

After saving, a ZIP-level restoration step re-injects any x14
extensions that openpyxl strips during its load / save cycle so the
workbook opens in Excel without corruption warnings.

Usage::

    from reporting.csa_workbook import build_csa_workbook
    build_csa_workbook(
        run_path="out/run.json",
        output_path="out/CSA_Workbook_v1.xlsm",
        why_payloads=[...],
    )
"""
from __future__ import annotations

import json
import os
import re
import shutil
import warnings
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# ══════════════════════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════════════════════

def _load_json(path: str | None) -> dict:
    if not path or not Path(path).exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _safe_get(obj: Any, dotpath: str, default: Any = "") -> Any:
    for key in dotpath.split("."):
        if isinstance(obj, dict):
            obj = obj.get(key, {})
        else:
            return default
    return obj if obj != {} else default


def _join_list(value) -> str:
    if isinstance(value, list):
        return "; ".join(str(v) for v in value if v)
    return str(value) if value else ""


# ══════════════════════════════════════════════════════════════════
# Template + sheet constants
# ══════════════════════════════════════════════════════════════════

_TEMPLATE_DIR = Path(__file__).resolve().parent
_TEMPLATE_NAME = "Landing_Zone_Assessment.xlsm"
_TEMPLATE_PATH = _TEMPLATE_DIR / _TEMPLATE_NAME

# Sheet names — must match template exactly
_SHEET_EXEC     = "0_Executive_Summary"
_SHEET_ROADMAP  = "1_30-60-90_Roadmap"
_SHEET_CONTROLS = "2_Control_Details"
_SHEET_RISK     = "3_Risk_Analysis"

# Control Details layout (row 9 = headers, row 10+ = data)
_CD_HEADER_ROW = 9
_CD_DATA_START = 10

_STATUS_MAP: dict[str, str] = {
    "Pass":         "Fulfilled",
    "Fail":         "Open",
    "Manual":       "Not verified",
    "Partial":      "Open",
    "Fulfilled":    "Fulfilled",
    "Open":         "Open",
    "Not verified": "Not verified",
    "Not required": "Not required",
    "N/A":          "N/A",
}


def _map_status(raw: str) -> str:
    return _STATUS_MAP.get(raw, "Not verified")


# ══════════════════════════════════════════════════════════════════
# Data clearing
# ══════════════════════════════════════════════════════════════════

def _clear_data_rows(ws, start_row: int = _CD_DATA_START, max_col: int = 25):
    """Clear data rows without touching headers or table structure."""
    from openpyxl.cell.cell import MergedCell
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = None


def _write_control_detail_rows(
    ws,
    results: list[dict],
    checklist_lookup: dict[str, dict],
) -> int:
    """Populate 2_Control_Details starting at row 10.

    Columns A–U:
      A: ID  B: Design Area  C: Sub Area  D: WAF Pillar  E: Service
      F: Checklist item  G: Description  H: Severity  I: Status
      J: Comment  K: AMMP  L: More info  M: Training  N: Graph Query
      O: GUID  P–T: (reserved)  U: Source File

    Returns the number of rows written.
    """
    row = _CD_DATA_START

    for ctrl in results:
        cid = ctrl.get("control_id", "")
        cl = checklist_lookup.get(cid, {})

        ws.cell(row=row, column=1,  value=cl.get("id", ""))
        ws.cell(row=row, column=2,  value=cl.get(
            "category", ctrl.get("category", ctrl.get("section", ""))))
        ws.cell(row=row, column=3,  value=cl.get("subcategory", ""))
        ws.cell(row=row, column=4,  value=cl.get("waf", ""))
        ws.cell(row=row, column=5,  value=cl.get("service", ""))
        ws.cell(row=row, column=6,  value=cl.get(
            "text", ctrl.get("text", ctrl.get("question", ""))))
        ws.cell(row=row, column=7,  value="")
        ws.cell(row=row, column=8,  value=ctrl.get(
            "severity", cl.get("severity", "")))
        ws.cell(row=row, column=9,  value=_map_status(
            ctrl.get("status", "Manual")))

        # Comment / evidence
        evidence = ctrl.get("evidence", [])
        parts: list[str] = []
        notes = ctrl.get("notes", "")
        if notes:
            parts.append(notes)
        for ev in evidence[:2]:
            if isinstance(ev, dict):
                s = ev.get("summary", ev.get("resource_id", ""))
                if s:
                    parts.append(str(s)[:120])
        ws.cell(row=row, column=10, value="\n".join(parts))

        ws.cell(row=row, column=11, value="")                       # AMMP
        ws.cell(row=row, column=12, value=cl.get("link", ""))       # Learn link
        ws.cell(row=row, column=13, value=cl.get("training", ""))   # Training
        ws.cell(row=row, column=14, value=ctrl.get("signal_used", ""))
        ws.cell(row=row, column=15, value=cid)                      # GUID
        for c in range(16, 21):
            ws.cell(row=row, column=c, value="")                    # P–T
        ws.cell(row=row, column=21, value="lz-assessor")            # Source

        row += 1

    return row - _CD_DATA_START


# ══════════════════════════════════════════════════════════════════
# 0_Executive_Summary  — populate existing rows
# ══════════════════════════════════════════════════════════════════

def _populate_executive_summary(ws, run: dict) -> None:
    """Write values into the existing Executive Summary layout.

    Template layout (column A = labels, column B = values):
      Row 1:  CSA ENGAGEMENT FRAMING (title — leave as-is)
      Row 2:  Engagement Objective | <text>
      Row 3:  Key Message | <text>
      Row 4:  Customer Outcome | <text>
      Row 6:  ASSESSMENT METRICS (title — leave as-is)
      Row 7:  Tenant ID | <value>
      Row 8:  Assessment Date | <value>
      Row 9:  Enterprise-Scale Ready | <value>
      Row 10: Overall Maturity | <value>
      Row 11: Data-Driven Controls | <count>
      Row 12: Requires Customer Input | <count>
      Row 13: Subscriptions Assessed | <count>
      Row 15: Top Risks (title — leave as-is)
      Row 16: Risk | Business Impact | Severity (sub-header — leave as-is)
      Row 17+: risk data rows
    """
    es = run.get("executive_summary", {})
    scoring = run.get("scoring", {})
    ec = run.get("execution_context", {})
    ai = run.get("ai", {})
    esr = ai.get("enterprise_scale_readiness", {})
    results = run.get("results", [])
    total_controls = len(results)
    top_risks = es.get("top_business_risks", [])

    # ── Engagement framing ────────────────────────────────────────
    risk_titles = [r.get("title", "") for r in top_risks[:5]]

    objective = (
        "Assess the customer's Azure landing zone maturity, identify "
        "critical gaps, and deliver a prioritised 30-60-90 remediation "
        "roadmap aligned to Microsoft Cloud Adoption Framework."
    )
    key_message = (
        f"This assessment evaluated {total_controls} controls across the "
        f"tenant using live platform telemetry. Top risk areas include: "
        f"{', '.join(risk_titles)}. The roadmap ties each action to "
        f"specific controls and risks, making every recommendation "
        f"defensible and auditable."
    )
    customer_outcome = (
        "A data-driven workbook the customer owns — with scored controls, "
        "a traceable remediation plan, and Microsoft Learn references — "
        "enabling them to drive implementation with or without further "
        "Microsoft engagement."
    )

    ws.cell(row=2, column=2, value=objective)
    ws.cell(row=3, column=2, value=key_message)
    ws.cell(row=4, column=2, value=customer_outcome)

    # ── Assessment metrics ────────────────────────────────────────
    ws.cell(row=7, column=2, value=ec.get("tenant_id", "Unknown"))
    ws.cell(row=8, column=2, value=run.get("meta", {}).get("timestamp", ""))

    ready = esr.get("ready_for_enterprise_scale", False)
    score = esr.get("readiness_score", "")
    ws.cell(row=9, column=2, value="Yes" if ready else f"No  (score: {score})")
    maturity = scoring.get('overall_maturity_percent')
    ws.cell(row=10, column=2,
            value=f"{maturity}%" if maturity is not None else "Unavailable")

    data_driven = sum(
        1 for r in results
        if r.get("status") in ("Pass", "Fail", "Partial", "Fulfilled", "Open")
        and r.get("signal_used")
    )
    ws.cell(row=11, column=2, value=data_driven)
    ws.cell(row=12, column=2, value=total_controls - data_driven)
    ws.cell(row=13, column=2, value=ec.get("subscription_count_visible", ""))

    # ── Top risks table (row 17+) ─────────────────────────────────
    row = 17
    for risk in top_risks:
        ws.cell(row=row, column=1, value=risk.get("title", ""))
        ws.cell(row=row, column=2, value=risk.get("business_impact", ""))
        ws.cell(row=row, column=3, value=risk.get("severity", ""))
        row += 1


# ══════════════════════════════════════════════════════════════════
# 1_30-60-90_Roadmap  — populate existing rows
# ══════════════════════════════════════════════════════════════════

def _populate_roadmap(ws, run: dict) -> int:
    """Write values into the existing Roadmap layout (row 1 = headers).

    Columns: Phase | Action | Initiative ID | CAF Discipline | Owner |
             Success Criteria | Dependencies | Related Controls | Related Risks

    Returns the number of rows written.
    """
    tr = run.get("transformation_roadmap", {})
    roadmap = tr.get("roadmap_30_60_90", {})
    tp = run.get("transformation_plan", {})
    init_lookup: dict[str, dict] = {
        i.get("initiative_id", ""): i
        for i in tp.get("initiatives", [])
        if i.get("initiative_id")
    }

    phase_map = {"30_days": "30 Days", "60_days": "60 Days", "90_days": "90 Days"}
    row = 2
    for phase_key, phase_label in phase_map.items():
        for item in roadmap.get(phase_key, []):
            iid = item.get("initiative_id", "")
            init_detail = init_lookup.get(iid, {})
            ws.cell(row=row, column=1, value=phase_label)
            ws.cell(row=row, column=2, value=item.get("action", ""))
            ws.cell(row=row, column=3, value=iid)
            ws.cell(row=row, column=4, value=item.get("caf_discipline", ""))
            ws.cell(row=row, column=5, value=item.get("owner_role", ""))
            ws.cell(row=row, column=6, value=item.get("success_criteria", ""))
            ws.cell(row=row, column=7, value=_join_list(
                item.get("dependency_on", [])))
            ws.cell(row=row, column=8, value=_join_list(
                init_detail.get("controls", [])))
            ws.cell(row=row, column=9, value="")  # filled by cross-ref below
            row += 1

    _cross_ref_roadmap_risks(ws, run, start_row=2, end_row=row - 1)
    return row - 2


def _cross_ref_roadmap_risks(
    ws, run: dict, start_row: int, end_row: int,
) -> None:
    """Fill column I (Related Risks) by matching initiative controls."""
    top_risks = run.get("executive_summary", {}).get("top_business_risks", [])
    for r in range(start_row, end_row + 1):
        related_ctrls = str(ws.cell(row=r, column=8).value or "")
        if not related_ctrls:
            continue
        ctrl_ids = {c.strip() for c in related_ctrls.replace(";", ",").split(",")}
        matched: list[str] = []
        for risk in top_risks:
            affected = risk.get("affected_controls", [])
            affected_shorts = {str(c)[:8] for c in affected}
            if ctrl_ids & affected_shorts:
                matched.append(risk.get("title", ""))
        if matched:
            ws.cell(row=r, column=9, value="; ".join(matched))


# ══════════════════════════════════════════════════════════════════
# 3_Risk_Analysis  — causal risk blocks
# ══════════════════════════════════════════════════════════════════

def _populate_risk_analysis(ws, why_payloads: list[dict]) -> int:
    """Write risk analysis blocks into the existing sheet.

    Each risk block follows the template pattern:
      Title → Root Cause → Business Impact → Failing Controls table →
      Dependency Impact table → Remediation Roadmap table →
      Cascade Effect → blank separator.

    Returns the number of risk blocks written.
    """
    if not why_payloads:
        return 0

    # Unmerge all cells first so we can write freely
    for merge in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge))

    _clear_data_rows(ws, start_row=1, max_col=7)

    row = 1
    for wp in why_payloads:
        risk = wp.get("risk", {})
        domain = wp.get("domain", "Unknown")
        title = risk.get("title", domain)

        # ── Title ─────────────────────────────────────────────────
        ws.cell(row=row, column=1,
                value=f"  {domain.upper()} — {title}")
        row += 1

        # ── Root Cause ────────────────────────────────────────────
        ws.cell(row=row, column=1, value="  Root Cause")
        row += 1
        ws.cell(row=row, column=1, value=risk.get("technical_cause", ""))
        row += 2

        # ── Business Impact ───────────────────────────────────────
        ws.cell(row=row, column=1, value="  Business Impact")
        row += 1
        ws.cell(row=row, column=1, value=risk.get("business_impact", ""))
        row += 2

        # ── Failing / Partial Controls ────────────────────────────
        ws.cell(row=row, column=1, value="  Failing / Partial Controls")
        row += 1
        for ci, h in enumerate(
            ["Control ID", "Section", "Severity", "Status",
             "Description", "Notes"], 1,
        ):
            ws.cell(row=row, column=ci, value=h)
        row += 1
        for fc in wp.get("failing_controls", []):
            ws.cell(row=row, column=1, value=str(fc.get("control_id", ""))[:8])
            ws.cell(row=row, column=2, value=fc.get("section", ""))
            ws.cell(row=row, column=3, value=fc.get("severity", ""))
            ws.cell(row=row, column=4, value=fc.get("status", ""))
            ws.cell(row=row, column=5, value=fc.get("text", ""))
            ws.cell(row=row, column=6, value=fc.get("notes", ""))
            row += 1
        row += 1

        # ── Dependency Impact ─────────────────────────────────────
        deps = wp.get("dependency_impact", [])
        if deps:
            ws.cell(row=row, column=1, value="  Dependency Impact")
            row += 1
            for ci, h in enumerate(
                ["Failing Control", "Name", "Blocks Count",
                 "Blocked Controls"], 1,
            ):
                ws.cell(row=row, column=ci, value=h)
            row += 1
            for dep in deps:
                ws.cell(row=row, column=1, value=str(dep.get("control", ""))[:8])
                ws.cell(row=row, column=2, value=dep.get("name", ""))
                ws.cell(row=row, column=3, value=str(dep.get("blocks_count", "")))
                blocks = dep.get("blocks", [])
                ws.cell(row=row, column=4,
                        value=", ".join(str(b)[:8] for b in blocks))
                row += 1
            row += 1

        # ── Remediation Roadmap ───────────────────────────────────
        ws.cell(row=row, column=1,
                value="  Remediation Roadmap (AI-Prioritized)")
        row += 1
        for ci, h in enumerate(
            ["Step", "Action", "Why This Order", "Phase", "Learn URL"], 1,
        ):
            ws.cell(row=row, column=ci, value=h)
        row += 1

        ai_explanation = wp.get("ai_explanation", {})
        ai_steps = (
            ai_explanation.get("remediation_steps", [])
            if isinstance(ai_explanation, dict) else []
        )
        steps = ai_steps or wp.get("roadmap_actions", [])
        for step_idx, step in enumerate(steps, 1):
            ws.cell(row=row, column=1, value=str(step_idx))
            ws.cell(row=row, column=2,
                    value=step.get("action", step.get("title", "")))
            ws.cell(row=row, column=3,
                    value=step.get("why_this_order", step.get("rationale", "")))
            ws.cell(row=row, column=4, value=step.get("phase", ""))
            refs = step.get("learn_references", [])
            url = ""
            if refs and isinstance(refs, list):
                first = refs[0]
                url = first.get("url", "") if isinstance(first, dict) else str(first)
            ws.cell(row=row, column=5, value=url)
            row += 1
        row += 1

        # ── Cascade Effect ────────────────────────────────────────
        ws.cell(row=row, column=1, value="  Cascade Effect")
        row += 1
        cascade = (
            ai_explanation.get("cascade_effect", "")
            if isinstance(ai_explanation, dict) else ""
        )
        if not cascade and deps:
            blocked_names = []
            for dep in deps:
                blocked_names.extend(str(b) for b in dep.get("blocks", []))
            if blocked_names:
                cascade = (
                    f"Remediating these root causes will unblock downstream "
                    f"controls: {', '.join(blocked_names[:5])}."
                )
        ws.cell(row=row, column=1, value=cascade)
        row += 3  # separator before next block

    return len(why_payloads)


# ══════════════════════════════════════════════════════════════════
# Signal integrity validation
# ══════════════════════════════════════════════════════════════════

class SignalIntegrityError(RuntimeError):
    """Raised when no platform signals were collected — report would be hollow."""


def validate_signal_integrity(run: dict, *, allow_demo: bool = False) -> dict:
    """Verify the run contains live platform signals before rendering.

    Returns a ``provenance`` dict with scan duration, API counts, and
    signal inventory.  Raises ``SignalIntegrityError`` if signal counts
    are zero and ``allow_demo`` is False.
    """
    telemetry = run.get("telemetry", {})
    is_live = telemetry.get("live_run", False)
    sig_avail = run.get("signal_availability", {})
    results = run.get("results", [])

    # Count signals that actually returned data — use None when absent
    rg_queries = telemetry.get("rg_query_count")
    arm_calls = telemetry.get("arm_call_count")
    signals_fetched = telemetry.get("signals_fetched")
    total_api_calls = (rg_queries or 0) + (arm_calls or 0)

    # Signal inventory from availability matrix
    signal_inventory: dict[str, int] = {}
    for category, sigs in sig_avail.items():
        if isinstance(sigs, list):
            signal_inventory[category] = len(sigs)

    # Data-driven controls (have a signal_used value)
    data_driven = sum(1 for r in results if r.get("signal_used"))

    provenance = {
        "live": is_live,
        "statement": (
            "This report was generated from live platform telemetry. "
            "No questionnaire or Excel input was used."
        ) if is_live else (
            "Demo Mode \u2014 No live telemetry. "
            "Metrics shown are from cached or sample data."
        ),
        "scan_duration_sec": telemetry.get("assessment_duration_sec"),
        "api_calls_total": total_api_calls if is_live else None,
        "rg_queries": rg_queries,
        "arm_calls": arm_calls,
        "signals_fetched": signals_fetched,
        "signals_cached": telemetry.get("signals_cached"),
        "signal_errors": telemetry.get("signal_errors"),
        "signal_inventory": signal_inventory,
        "signal_categories": len(signal_inventory),
        "data_driven_controls": data_driven,
        "total_controls": len(results),
    }

    # Gate: abort if no live signals and not demo
    if total_api_calls == 0 and data_driven == 0 and not allow_demo:
        raise SignalIntegrityError(
            "ABORT: Platform signal counts are zero. "
            "No live telemetry was collected — cannot generate a credible report. "
            f"(rg_queries={rg_queries}, arm_calls={arm_calls}, "
            f"data_driven_controls={data_driven})"
        )

    return provenance


# ══════════════════════════════════════════════════════════════════
# ZIP-level extension restoration
# ══════════════════════════════════════════════════════════════════

_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_REL  = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _zip_sheet_map(zf: zipfile.ZipFile) -> dict[str, str]:
    """Return ``{sheet_name: zip_path}`` from workbook.xml + rels."""
    wb = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rid_target: dict[str, str] = {}
    for r in rels:
        t = r.get("Target", "")
        # Normalise to ZIP-entry path (no leading /, relative to xl/)
        t = t.lstrip("/")
        if not t.startswith("xl/"):
            t = "xl/" + t
        rid = r.get("Id", "")
        if rid:
            rid_target[rid] = t

    result: dict[str, str] = {}
    for sh in wb.findall(f".//{{{_NS_MAIN}}}sheet"):
        name = sh.get("name", "")
        rid = sh.get(f"{{{_NS_REL}}}id")
        if name and rid and rid in rid_target:
            result[name] = rid_target[rid]
    return result


def _extract_ws_extlst(data: bytes) -> bytes | None:
    """Extract the worksheet-level ``<extLst>…</extLst>`` block.

    The worksheet-level extLst is always the last child element before
    ``</worksheet>``.  Handles nested extLst elements correctly.
    """
    ws_end = data.rfind(b"</worksheet>")
    if ws_end == -1:
        return None
    region = data[:ws_end]

    start = region.rfind(b"<extLst")
    if start == -1:
        return None

    # Walk forward to find the matching </extLst> (handles nesting)
    depth = 0
    pos = start
    OPEN = b"<extLst"
    CLOSE = b"</extLst>"
    # Move past the initial tag
    scan_from = start + len(OPEN)
    while scan_from < ws_end:
        next_open = data.find(OPEN, scan_from)
        next_close = data.find(CLOSE, scan_from)

        if next_close == -1:
            return None

        if next_open != -1 and next_open < next_close:
            depth += 1
            scan_from = next_open + len(OPEN)
        else:
            if depth == 0:
                return data[start : next_close + len(CLOSE)]
            depth -= 1
            scan_from = next_close + len(CLOSE)
    return None


def _extract_ns_decls(data: bytes) -> list[bytes]:
    """Extract ``xmlns:*`` declarations from the root element tag."""
    # Skip XML declaration if present
    if data.startswith(b"<?"):
        decl_end = data.find(b"?>")
        root_start = data.find(b"<", decl_end + 2)
    else:
        root_start = 0
    root_end = data.find(b">", root_start)
    if root_end == -1:
        return []
    root_tag = data[root_start:root_end]
    return re.findall(rb'xmlns:\w+="[^"]*"', root_tag)


def _extract_mc_ignorable(data: bytes) -> bytes | None:
    """Extract ``mc:Ignorable`` attribute value from the root tag."""
    if data.startswith(b"<?"):
        decl_end = data.find(b"?>")
        root_start = data.find(b"<", decl_end + 2)
    else:
        root_start = 0
    root_end = data.find(b">", root_start)
    if root_end == -1:
        return None
    root_tag = data[root_start:root_end]
    m = re.search(rb'mc:Ignorable="([^"]*)"', root_tag)
    return m.group(1) if m else None


def _root_tag_end(data: bytes) -> int:
    """Return the byte offset of the first ``>`` in the root element."""
    if data.startswith(b"<?"):
        decl_end = data.find(b"?>")
        root_start = data.find(b"<", decl_end + 2)
    else:
        root_start = 0
    return data.find(b">", root_start)


def _restore_extensions(template_path: str, output_path: str) -> int:
    """Re-inject x14 extensions that openpyxl strips on load / save.

    For every worksheet present in **both** the template and the output,
    copies the ``<extLst>`` block (x14 conditional formatting, data
    validation, etc.) from the template back into the output.  Also
    ensures the required ``xmlns:*`` and ``mc:Ignorable`` declarations
    exist on the ``<worksheet>`` root element.

    The output ZIP is rewritten in-place (via a temp file).
    Returns the number of sheets patched.
    """
    # ── Read template sheet data ──────────────────────────────────
    with zipfile.ZipFile(template_path, "r") as zt:
        tpl_map = _zip_sheet_map(zt)
        tpl_data: dict[str, bytes] = {}
        for name, path in tpl_map.items():
            try:
                tpl_data[name] = zt.read(path)
            except KeyError:
                pass

    # ── Read all output ZIP entries (preserving order) ────────────
    with zipfile.ZipFile(output_path, "r") as zo:
        out_map = _zip_sheet_map(zo)
        entry_order = zo.namelist()
        entries: dict[str, bytes] = {n: zo.read(n) for n in entry_order}

    # ── Patch each sheet ──────────────────────────────────────────
    patched = 0
    for sheet_name, tpl_bytes in tpl_data.items():
        if sheet_name not in out_map or sheet_name == "ARG":
            continue
        out_path = out_map[sheet_name]
        if out_path not in entries:
            continue

        extlst = _extract_ws_extlst(tpl_bytes)
        if extlst is None:
            continue

        out_bytes = entries[out_path]

        # Remove any partial extLst openpyxl may have left
        existing = _extract_ws_extlst(out_bytes)
        if existing:
            out_bytes = out_bytes.replace(existing, b"")

        # Inject template extLst before </worksheet>
        ws_end = out_bytes.rfind(b"</worksheet>")
        out_bytes = out_bytes[:ws_end] + extlst + b"\n" + out_bytes[ws_end:]

        # Ensure namespace declarations from template are present
        tpl_ns = _extract_ns_decls(tpl_bytes)
        for ns_decl in tpl_ns:
            rte = _root_tag_end(out_bytes)
            if ns_decl not in out_bytes[:rte]:
                out_bytes = out_bytes[:rte] + b" " + ns_decl + out_bytes[rte:]

        # Ensure mc:Ignorable matches the template
        tpl_mc = _extract_mc_ignorable(tpl_bytes)
        if tpl_mc:
            out_mc = _extract_mc_ignorable(out_bytes)
            if out_mc != tpl_mc:
                rte = _root_tag_end(out_bytes)
                root_tag = out_bytes[:rte]
                if out_mc:
                    root_tag = root_tag.replace(
                        b'mc:Ignorable="' + out_mc + b'"',
                        b'mc:Ignorable="' + tpl_mc + b'"',
                    )
                else:
                    root_tag += b' mc:Ignorable="' + tpl_mc + b'"'
                out_bytes = root_tag + out_bytes[rte:]

        entries[out_path] = out_bytes
        patched += 1

    # ── Rewrite the ZIP ───────────────────────────────────────────
    if patched > 0:
        tmp = output_path + ".tmp"
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for name in entry_order:
                if name in entries:
                    zout.writestr(name, entries[name])
        os.replace(tmp, output_path)

    return patched


# ══════════════════════════════════════════════════════════════════
# Main builder
# ══════════════════════════════════════════════════════════════════

def build_csa_workbook(
    run_path: str = "out/run.json",
    target_path: str = "out/target_architecture.json",
    output_path: str = "out/CSA_Workbook_v1.xlsm",
    why_payloads: list[dict] | None = None,
    template_path: str | None = None,
) -> str:
    """Build the CSA workbook: copy template, populate all 4 data sheets.

    This function is a **data injector** — it writes only values into the
    template's existing data ranges.  It performs NO scoring, NO inference,
    NO sheet creation/deletion, and NO formatting changes.

    Sheets populated:
      ``0_Executive_Summary`` — engagement framing + metrics + top risks
      ``1_30-60-90_Roadmap`` — phased initiatives
      ``2_Control_Details`` — one row per control (A–U) + enrichment (V–Y)
      ``3_Risk_Analysis`` — causal risk blocks from why-analysis
    """
    run = _load_json(run_path)

    # ── Resolve template ──────────────────────────────────────────
    tpl = Path(template_path) if template_path else _TEMPLATE_PATH
    if not tpl.exists():
        raise FileNotFoundError(
            f"Template not found: {tpl}\n"
            f"Place {_TEMPLATE_NAME} in {_TEMPLATE_DIR}/"
        )

    # ── Log provenance before rendering ───────────────────────────
    ec = run.get("execution_context", {})
    telem = run.get("telemetry", {})
    print("  ┌─ Workbook Provenance ────────────────┐")
    print(f"  │ tenant_id:              {ec.get('tenant_id', 'N/A')}")
    print(f"  │ subscription_count:     {ec.get('subscription_count_visible', 'N/A')}")
    print(f"  │ rg_queries:             {telem.get('rg_query_count', 0)}")
    print(f"  │ arm_calls:              {telem.get('arm_call_count', 0)}")
    print(f"  │ signals_fetched:        {telem.get('signals_fetched', 0)}")
    print(f"  │ scan_duration:          {telem.get('assessment_duration_sec', 0)}s")
    print("  └─────────────────────────────────────────┘")

    # ── Copy template → output (byte-for-byte) ───────────────────
    out = Path(output_path)
    if out.suffix.lower() != ".xlsm":
        out = out.with_suffix(".xlsm")
    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(str(tpl), str(out))

    # ── Single openpyxl pass (suppress extension warnings) ────────
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(str(out), keep_vba=True)

    # NEVER delete sheets — template owns all structure
    # Verify expected sheets exist
    for sheet_name in [_SHEET_EXEC, _SHEET_ROADMAP, _SHEET_CONTROLS, _SHEET_RISK]:
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠ Sheet '{sheet_name}' not found in template — skipping")

    # ── Load ALZ checklist for rich per-control fields ────────────
    checklist_lookup: dict[str, dict] = {}
    try:
        from alz.loader import load_alz_checklist
        cl = load_alz_checklist()
        for item in cl.get("items", []):
            guid = item.get("guid", "")
            if guid:
                checklist_lookup[guid] = item
    except Exception:
        pass

    results = run.get("results", [])

    # ── 0_Executive_Summary ───────────────────────────────────────
    if _SHEET_EXEC in wb.sheetnames:
        _populate_executive_summary(wb[_SHEET_EXEC], run)
        print("  ✓ 0_Executive_Summary populated")

    # ── 1_30-60-90_Roadmap ────────────────────────────────────────
    if _SHEET_ROADMAP in wb.sheetnames:
        ws_rm = wb[_SHEET_ROADMAP]
        _clear_data_rows(ws_rm, start_row=2, max_col=9)
        n_roadmap = _populate_roadmap(ws_rm, run)
        print(f"  ✓ 1_30-60-90_Roadmap: {n_roadmap} initiatives")

    # ── 2_Control_Details (primary data sheet) ────────────────────
    if _SHEET_CONTROLS in wb.sheetnames:
        ws_cd = wb[_SHEET_CONTROLS]
        _clear_data_rows(ws_cd, start_row=_CD_DATA_START)
        n_controls = _write_control_detail_rows(ws_cd, results, checklist_lookup)

        # Enrichment in the same open workbook (no second load/save)
        try:
            from reporting.enrich import enrich_open_worksheet
            e_stats = enrich_open_worksheet(ws_cd)
            print(
                f"  ✓ 2_Control_Details: {n_controls} controls "
                f"({e_stats.get('alz', 0)} ALZ, "
                f"{e_stats.get('derived', 0)} derived)"
            )
        except Exception as e:
            print(f"  ✓ 2_Control_Details: {n_controls} controls "
                  f"(enrichment skipped: {e})")
    else:
        print(f"  ⚠ Sheet '{_SHEET_CONTROLS}' not found — skipping")

    # ── 3_Risk_Analysis ───────────────────────────────────────────
    if _SHEET_RISK in wb.sheetnames and why_payloads:
        n_risks = _populate_risk_analysis(wb[_SHEET_RISK], why_payloads)
        print(f"  ✓ 3_Risk_Analysis: {n_risks} risk blocks")
    elif _SHEET_RISK in wb.sheetnames:
        print("  ⚠ 3_Risk_Analysis: no why_payloads — keeping template data")

    # ── Save ──────────────────────────────────────────────────────
    try:
        wb.save(str(out))
    except PermissionError:
        ts = datetime.now().strftime("%H%M%S")
        fallback = out.with_name(f"{out.stem}_{ts}.xlsm")
        wb.save(str(fallback))
        print(f"  ⚠ Saved as {fallback.name} (original locked)")
        out = fallback

    # ── Restore x14 extensions stripped by openpyxl ───────────────
    patched = _restore_extensions(str(tpl), str(out))
    if patched:
        print(f"  ✓ Extensions restored for {patched} sheet(s)")

    print(f"  ✓ CSA workbook → {out}")
    return str(out)
